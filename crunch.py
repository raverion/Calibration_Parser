import os
import json
import pandas as pd
import numpy as np
from pathlib import Path
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.axis import ChartLines

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False


def parse_filename(filename):
    """
    Extract test value, unit, channel number (if present), and range setting from filename.
    Supports formats like:
    - VT2816A_m2V5_R10V_CH3.csv (voltage: -2.5V, range: 10V, channel: 3)
    - VIO2004_3mA_R10mA_CH1.txt (current: 3mA, range: 10mA, channel: 1)
    - VT2816A_10V_R10V_1000x.txt (voltage: 10V, range: 10V, no channel - multi-channel file)
    - VT2516A_25V_1000x.txt (voltage: 25V, no range, no channel - multi-channel file)
    """
    name = Path(filename).stem
    
    # Extract channel pattern (e.g., CH1, CH2, CH3, CH4) - may not be present
    channel_pattern = r'_CH(\d+)'
    channel_match = re.search(channel_pattern, name, re.IGNORECASE)
    channel_num = int(channel_match.group(1)) if channel_match else None
    
    # Extract range setting pattern (e.g., R10V, R10mA, R100ohm)
    range_pattern = r'_R(\d+(?:\.\d+)?)(V|mV|mA|uA|A|ohm|Ohm|kOhm|MOhm)(?:_|$)'
    range_match = re.search(range_pattern, name, re.IGNORECASE)
    
    range_setting = None
    if range_match:
        range_value = range_match.group(1)
        range_unit = range_match.group(2)
        range_setting = f"{range_value}{range_unit}"
    
    # Try voltage pattern first (m2V5, p7V5, 0V, 10V, 25V) - but not matching the R prefix range
    voltage_pattern = r'(?<!R)_([mp]?\d+V\d*)(?:_|$)'
    voltage_match = re.search(voltage_pattern, name, re.IGNORECASE)
    
    if voltage_match:
        voltage_str = voltage_match.group(1).lower()
        sign = -1 if voltage_str.startswith('m') else 1
        voltage_str = voltage_str.lstrip('mp').replace('v', '.')
        # Handle cases like "10V" -> "10." -> need to strip trailing dot
        if voltage_str.endswith('.'):
            voltage_str = voltage_str[:-1]
        try:
            value = sign * float(voltage_str)
            return value, 'V', channel_num, range_setting
        except ValueError:
            pass
    
    # Try milliampere pattern (e.g., 3mA, m5mA, p10mA)
    ma_pattern = r'(?<!R)_([mp]?\d+(?:\.\d+)?)\s*mA(?:_|$)'
    ma_match = re.search(ma_pattern, name, re.IGNORECASE)
    
    if ma_match:
        value_str = ma_match.group(1)
        sign = -1 if value_str.startswith('m') else 1
        value_str = value_str.lstrip('mp')
        try:
            value = sign * float(value_str)
            return value, 'mA', channel_num, range_setting
        except ValueError:
            pass
    
    # Try microampere pattern (e.g., 100uA, m50uA)
    ua_pattern = r'(?<!R)_([mp]?\d+(?:\.\d+)?)\s*uA(?:_|$)'
    ua_match = re.search(ua_pattern, name, re.IGNORECASE)
    
    if ua_match:
        value_str = ua_match.group(1)
        sign = -1 if value_str.startswith('m') else 1
        value_str = value_str.lstrip('mp')
        try:
            value = sign * float(value_str)
            return value, 'uA', channel_num, range_setting
        except ValueError:
            pass
    
    # Try ampere pattern (e.g., 1A, 2A)
    a_pattern = r'(?<!R|m|u)_([mp]?\d+(?:\.\d+)?)\s*A(?:_|$)'
    a_match = re.search(a_pattern, name, re.IGNORECASE)
    
    if a_match:
        value_str = a_match.group(1)
        sign = -1 if value_str.startswith('m') else 1
        value_str = value_str.lstrip('mp')
        try:
            value = sign * float(value_str)
            return value, 'A', channel_num, range_setting
        except ValueError:
            pass
    
    # Try ohms pattern (10_ohms, 100ohms, etc.)
    ohms_pattern = r'_(\d+(?:\.\d+)?)[_\s]?ohms?(?:_|$)'
    ohms_match = re.search(ohms_pattern, name, re.IGNORECASE)
    
    if ohms_match:
        try:
            value = float(ohms_match.group(1))
            return value, 'Ohm', channel_num, range_setting
        except ValueError:
            pass
    
    # Try generic numeric pattern with underscore
    generic_pattern = r'_([mp]?\d+(?:\.\d+)?)_'
    generic_match = re.search(generic_pattern, name)
    
    if generic_match:
        value_str = generic_match.group(1)
        sign = -1 if value_str.startswith('m') else 1
        value_str = value_str.lstrip('mp')
        try:
            value = sign * float(value_str)
            return value, 'unknown', channel_num, range_setting
        except ValueError:
            pass
    
    return None, None, channel_num, range_setting


def get_unit_from_files(input_dir):
    """
    Determine the measurement unit from filenames.
    """
    all_files = list(Path(input_dir).glob('*.csv')) + list(Path(input_dir).glob('*.txt'))
    
    for file in all_files:
        _, unit, _, _ = parse_filename(file.name)
        if unit and unit != 'unknown':
            return unit
    
    return 'V'  # Default to volts


def scan_text_file_for_measurement_types(file_path):
    """
    Scan a text file to find all unique measurement types per channel.
    Returns a set of measurement type names found (e.g., {'Voltage', 'MeanVoltage'} or {'CurVoltage'} or {'Avg'})
    """
    measurement_types = set()
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except Exception:
        return measurement_types
    
    lines = content.strip().split('\n')
    
    # Check for hierarchical format (VIO1008 style)
    # Pattern: |  MeasurementType_Chxx   value   unit   ...
    hierarchical_pattern = r'\|\s+(\w+)_Ch\d+'
    for line in lines[:500]:  # Check first 500 lines
        match = re.search(hierarchical_pattern, line)
        if match:
            measurement_types.add(match.group(1))
    
    if measurement_types:
        return measurement_types
    
    # Check for flat format (VT2816A/VT2516A style)
    # Pattern: Time  Name::MeasurementType  Data
    flat_pattern = r'_Ch\d+::(\w+)'
    for line in lines[:500]:
        match = re.search(flat_pattern, line, re.IGNORECASE)
        if match:
            measurement_types.add(match.group(1))
    
    return measurement_types


def parse_text_file(file_path, selected_measurement_type=None, channel_from_filename=None):
    """
    Parse text files containing measurement data from multiple channels.
    
    Supports multiple formats:
    1. Hierarchical format (VIO1008 style):
       [-] timestamp   TaskName
             |  Voltage_Ch01       -2.498169   V   ...
             |  MeanVoltage_Ch01   -2.498347   V   ...
    
    2. Flat format with channel in name (VT2816A style):
       Time        Name                        Data
       66.001210   VT2816_1_Ch1::CurVoltage    10.011883
    
    3. Flat format with channel in name (VT2516A style):
       Time        Name                 Data
       30.000132   VT2516_1_Ch1::Avg    24.976000
    
    4. Simple flat format without channel in data (VN1630A style):
       Time        Name            Data
       15.001821   VN1600_1::AIN   0.686400
       (Channel comes from filename, e.g., VN1630A_0V7_CH1_100x.txt)
    
    Parameters:
    - file_path: Path to the text file
    - selected_measurement_type: If file has multiple measurement types per channel,
                                 use this one (e.g., 'Voltage' or 'MeanVoltage')
    - channel_from_filename: Channel number parsed from filename (used for format 4)
    
    Returns: Dictionary mapping channel numbers to lists of measurements
    """
    channel_data = {}
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return channel_data
    
    lines = content.strip().split('\n')
    if not lines:
        return channel_data
    
    # Try hierarchical format first (VIO1008 style)
    # Pattern: |  MeasurementType_Chxx   value   unit   value   description
    hierarchical_pattern = r'\|\s+(\w+)_Ch(\d+)\s+(-?\d+\.?\d*)\s+(\w+)'
    
    hierarchical_matches = []
    for line in lines:
        match = re.search(hierarchical_pattern, line)
        if match:
            hierarchical_matches.append({
                'type': match.group(1),
                'channel': int(match.group(2)),
                'value': float(match.group(3))
            })
    
    if hierarchical_matches:
        # Filter by selected measurement type if specified
        for match in hierarchical_matches:
            if selected_measurement_type and match['type'] != selected_measurement_type:
                continue
            
            channel = match['channel']
            value = match['value']
            
            if channel not in channel_data:
                channel_data[channel] = []
            channel_data[channel].append(value)
        
        if channel_data:
            return channel_data
    
    # Try flat format with channel in name (VT2816A/VT2516A style)
    # Pattern: Time  DeviceName_Chxx::MeasurementType  Data
    flat_pattern = r'^\s*[\d.]+\s+\S+_Ch(\d+)::(\w+)\s+(-?\d+\.?\d*)'
    
    for line in lines:
        match = re.search(flat_pattern, line)
        if match:
            channel = int(match.group(1))
            meas_type = match.group(2)
            value = float(match.group(3))
            
            # Filter by selected measurement type if specified
            if selected_measurement_type and meas_type != selected_measurement_type:
                continue
            
            if channel not in channel_data:
                channel_data[channel] = []
            channel_data[channel].append(value)
    
    if channel_data:
        return channel_data
    
    # Try simple flat format without channel in data (VN1630A style)
    # Pattern: Time  Name::Type  Data  OR  Time  Name  Data
    # Channel must come from filename
    simple_flat_pattern = r'^\s*[\d.]+\s+\S+\s+(-?\d+\.?\d*)\s*$'
    
    values_found = []
    for line in lines:
        match = re.search(simple_flat_pattern, line)
        if match:
            try:
                value = float(match.group(1))
                values_found.append(value)
            except ValueError:
                continue
    
    if values_found:
        # Use channel from filename if provided, otherwise default to channel 1
        channel = channel_from_filename if channel_from_filename is not None else 1
        channel_data[channel] = values_found
        return channel_data
    
    return channel_data


def select_measurement_type(file_measurement_types):
    """
    Show a dialog to let user select which measurement type to process
    when files have multiple measurement types per channel.
    
    Parameters:
    - file_measurement_types: Dict mapping filename to set of measurement types
    
    Returns:
    - Dict mapping filename to selected measurement type, or None if cancelled
    """
    # Check if any file has multiple measurement types
    files_with_multiple = {f: types for f, types in file_measurement_types.items() if len(types) > 1}
    
    if not files_with_multiple:
        # No selection needed - return the single type for each file
        return {f: list(types)[0] if types else None for f, types in file_measurement_types.items()}
    
    # Create selection dialog
    root = tk.Tk()
    root.title("Measurement Type Selection")
    root.geometry("700x500")
    root.resizable(True, True)
    
    style = ttk.Style()
    style.theme_use('clam')
    
    # Header
    header_frame = tk.Frame(root, bg='#1F4E78', height=80)
    header_frame.pack(fill='x')
    header_frame.pack_propagate(False)
    
    title_label = tk.Label(header_frame, text="Measurement Type Selection", 
                          font=('Segoe UI', 14, 'bold'), 
                          bg='#1F4E78', fg='white')
    title_label.pack(pady=10)
    
    subtitle_label = tk.Label(header_frame, 
                             text="Some files have multiple measurement types per channel. Select which to process:", 
                             font=('Segoe UI', 10), 
                             bg='#1F4E78', fg='white')
    subtitle_label.pack()
    
    # Scrollable frame
    canvas = tk.Canvas(root, bg='white')
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='white')
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Selection widgets
    selection_vars = {}
    
    for i, (filename, types) in enumerate(files_with_multiple.items()):
        bg_color = '#F8F8F8' if i % 2 == 0 else 'white'
        frame = tk.Frame(scrollable_frame, bg=bg_color, pady=10)
        frame.pack(fill='x', padx=20, pady=5)
        
        # File label
        file_label = tk.Label(frame, text=Path(filename).name, 
                             font=('Segoe UI', 10, 'bold'), bg=bg_color, 
                             anchor='w', width=40)
        file_label.pack(side='left', padx=10)
        
        # Dropdown for measurement type
        types_list = sorted(list(types))
        var = tk.StringVar(value=types_list[0])
        selection_vars[filename] = var
        
        dropdown = ttk.Combobox(frame, textvariable=var, values=types_list, 
                               state='readonly', width=20)
        dropdown.pack(side='left', padx=10)
        
        # Description of types
        desc_label = tk.Label(frame, text=f"Available: {', '.join(types_list)}", 
                             font=('Segoe UI', 9), bg=bg_color, fg='gray')
        desc_label.pack(side='left', padx=10)
    
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=(10, 90))
    scrollbar.pack(side="right", fill="y", pady=(10, 90))
    
    # Button frame
    button_frame = tk.Frame(root, bg='#F0F0F0', height=80)
    button_frame.pack(fill='x', side='bottom', pady=0, before=canvas)
    button_frame.pack_propagate(False)
    button_frame.lift()
    
    result = {'cancelled': False}
    
    def on_submit():
        result['cancelled'] = False
        root.quit()
        root.destroy()
    
    def on_cancel():
        result['cancelled'] = True
        root.quit()
        root.destroy()
    
    submit_btn = tk.Button(button_frame, text="Continue", command=on_submit,
                          font=('Segoe UI', 11, 'bold'), bg='#0070C0', fg='white',
                          width=12, height=2, cursor='hand2', relief='raised', bd=2)
    submit_btn.pack(side='right', padx=20, pady=15)
    
    cancel_btn = tk.Button(button_frame, text="Cancel", command=on_cancel,
                          font=('Segoe UI', 11), bg='#E0E0E0', fg='black',
                          width=12, height=2, cursor='hand2', relief='raised', bd=2)
    cancel_btn.pack(side='right', padx=5, pady=15)
    
    # Center window
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f'+{x}+{y}')
    
    root.mainloop()
    
    if result['cancelled']:
        return None
    
    # Build result dictionary
    selections = {}
    for filename, types in file_measurement_types.items():
        if filename in selection_vars:
            selections[filename] = selection_vars[filename].get()
        elif len(types) == 1:
            selections[filename] = list(types)[0]
        else:
            selections[filename] = None
    
    return selections


def get_user_inputs(test_value_range_io_tuples, unit, input_dir=None):
    """
    Ask user to input range setting, reference value, and tolerance for each 
    test value/range/IO-type combination using a GUI.
    
    Parameters:
    - test_value_range_io_tuples: Set of (test_value, range_setting, io_type) tuples
    - unit: Measurement unit
    - input_dir: Directory path for saving/loading config files
    
    Returns:
    - Dictionary mapping (test_value, range_setting, io_type) to config dict
    """
    user_inputs = {}
    
    root = tk.Tk()
    root.title("Test Configuration Input")
    root.geometry("950x750")
    root.resizable(True, True)
    
    style = ttk.Style()
    style.theme_use('clam')
    
    # Header
    header_frame = tk.Frame(root, bg='#1F4E78', height=100)
    header_frame.pack(fill='x')
    header_frame.pack_propagate(False)
    
    title_label = tk.Label(header_frame, text="Test Configuration Input", 
                          font=('Segoe UI', 16, 'bold'), 
                          bg='#1F4E78', fg='white')
    title_label.pack(pady=10)
    
    subtitle_label = tk.Label(header_frame, 
                             text=f"Configure range, reference value, and tolerance for each test ({unit})", 
                             font=('Segoe UI', 10), 
                             bg='#1F4E78', fg='white')
    subtitle_label.pack()
    
    # Scrollable frame
    canvas = tk.Canvas(root, bg='white')
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = tk.Frame(canvas, bg='white')
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    # Column headers
    header_row = tk.Frame(scrollable_frame, bg='#E8E8E8', pady=8)
    header_row.pack(fill='x', padx=10, pady=(10, 5))
    
    tk.Label(header_row, text="Test Value", font=('Segoe UI', 10, 'bold'), 
             bg='#E8E8E8', width=14, anchor='w').pack(side='left', padx=5)
    tk.Label(header_row, text="I/O Type", font=('Segoe UI', 10, 'bold'), 
             bg='#E8E8E8', width=10, anchor='w').pack(side='left', padx=5)
    tk.Label(header_row, text="Range Setting", font=('Segoe UI', 10, 'bold'), 
             bg='#E8E8E8', width=12, anchor='w').pack(side='left', padx=5)
    tk.Label(header_row, text="Reference Value", font=('Segoe UI', 10, 'bold'), 
             bg='#E8E8E8', width=14, anchor='w').pack(side='left', padx=5)
    tk.Label(header_row, text="Tolerance (±)", font=('Segoe UI', 10, 'bold'), 
             bg='#E8E8E8', width=12, anchor='w').pack(side='left', padx=5)
    
    # Entry fields
    entry_widgets = {}
    # Sort by test value, then by I/O type (Input first, then Output), then by range
    sorted_tuples = sorted(test_value_range_io_tuples, key=lambda x: (x[0], x[2], x[1] or ''))
    
    for i, (test_value, range_setting, io_type) in enumerate(sorted_tuples):
        # Use different background colors for Input vs Output
        if io_type == 'Input':
            bg_color = '#E8F4E8' if i % 2 == 0 else '#F0FAF0'  # Light green tint
        else:
            bg_color = '#E8E8F4' if i % 2 == 0 else '#F0F0FA'  # Light blue tint
        
        frame = tk.Frame(scrollable_frame, bg=bg_color, pady=8)
        frame.pack(fill='x', padx=10, pady=2)
        
        # Test value label
        label = tk.Label(frame, text=f"{test_value:+.4g} {unit}", 
                        font=('Segoe UI', 10), bg=bg_color, width=14, anchor='w')
        label.pack(side='left', padx=5)
        
        # I/O Type label with color coding
        io_color = '#006400' if io_type == 'Input' else '#00008B'  # Dark green for Input, Dark blue for Output
        io_label = tk.Label(frame, text=io_type, 
                           font=('Segoe UI', 10, 'bold'), bg=bg_color, fg=io_color, width=10, anchor='w')
        io_label.pack(side='left', padx=5)
        
        # Range setting entry
        range_entry = ttk.Entry(frame, font=('Segoe UI', 10), width=12)
        range_entry.pack(side='left', padx=5)
        range_entry.insert(0, range_setting if range_setting else "N/A")
        
        # Reference value entry
        ref_entry = ttk.Entry(frame, font=('Segoe UI', 10), width=14)
        ref_entry.pack(side='left', padx=5)
        ref_entry.insert(0, f"{test_value:.6g}")
        
        # Tolerance entry
        tol_entry = ttk.Entry(frame, font=('Segoe UI', 10), width=12)
        tol_entry.pack(side='left', padx=5)
        tol_entry.insert(0, "0.015")
        
        # Unit label
        unit_label = tk.Label(frame, text=unit, font=('Segoe UI', 10), bg=bg_color, width=5)
        unit_label.pack(side='left')
        
        entry_widgets[(test_value, range_setting, io_type)] = {
            'range': range_entry,
            'reference': ref_entry,
            'tolerance': tol_entry
        }
    
    canvas.pack(side="left", fill="both", expand=True, padx=10, pady=(10, 140))
    scrollbar.pack(side="right", fill="y", pady=(10, 140))
    
    # Legend frame
    legend_frame = tk.Frame(scrollable_frame, bg='white', pady=10)
    legend_frame.pack(fill='x', padx=10, pady=(20, 5))
    
    tk.Label(legend_frame, text="Legend: ", font=('Segoe UI', 9, 'bold'), bg='white').pack(side='left', padx=5)
    tk.Label(legend_frame, text="■ Input (TXT files - e.g., Voltmeter readings)", 
             font=('Segoe UI', 9), bg='white', fg='#006400').pack(side='left', padx=10)
    tk.Label(legend_frame, text="■ Output (CSV files - e.g., Power supply output)", 
             font=('Segoe UI', 9), bg='white', fg='#00008B').pack(side='left', padx=10)
    
    # Button frame with two rows
    button_frame = tk.Frame(root, bg='#F0F0F0', height=130)
    button_frame.pack(fill='x', side='bottom', pady=0, before=canvas)
    button_frame.pack_propagate(False)
    button_frame.lift()
    
    # Config buttons row (Save/Load)
    config_row = tk.Frame(button_frame, bg='#F0F0F0')
    config_row.pack(fill='x', pady=(10, 5))
    
    # Status label for showing load/save messages
    status_var = tk.StringVar(value="")
    status_label = tk.Label(config_row, textvariable=status_var, 
                           font=('Segoe UI', 9, 'italic'), bg='#F0F0F0', fg='#666666')
    status_label.pack(side='left', padx=20)
    
    def save_config():
        """Save current configuration to a JSON file"""
        try:
            config_data = {
                'unit': unit,
                'configurations': []
            }
            
            for (test_value, range_setting, io_type), entries in entry_widgets.items():
                config_entry = {
                    'test_value': test_value,
                    'range_setting': range_setting,
                    'io_type': io_type,
                    'range_input': entries['range'].get().strip(),
                    'reference': entries['reference'].get().strip(),
                    'tolerance': entries['tolerance'].get().strip()
                }
                config_data['configurations'].append(config_entry)
            
            # Ask user for save location
            default_filename = "test_config.json"
            if input_dir:
                default_path = os.path.join(input_dir, default_filename)
            else:
                default_path = default_filename
            
            file_path = filedialog.asksaveasfilename(
                title="Save Configuration",
                initialdir=input_dir or os.getcwd(),
                initialfile=default_filename,
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(config_data, f, indent=2)
                status_var.set(f"✓ Configuration saved to {Path(file_path).name}")
                root.after(5000, lambda: status_var.set(""))  # Clear after 5 seconds
        except Exception as e:
            messagebox.showerror("Save Error", f"Failed to save configuration:\n{str(e)}")
    
    def load_config():
        """Load configuration from a JSON file"""
        try:
            file_path = filedialog.askopenfilename(
                title="Load Configuration",
                initialdir=input_dir or os.getcwd(),
                defaultextension=".json",
                filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
            )
            
            if file_path:
                apply_config_file(file_path)
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load configuration:\n{str(e)}")
    
    def apply_config_file(file_path):
        """Apply configuration from a file to the entry widgets"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
            
            # Check unit compatibility
            if config_data.get('unit') != unit:
                result = messagebox.askyesno(
                    "Unit Mismatch", 
                    f"Config file unit ({config_data.get('unit')}) differs from current unit ({unit}).\n"
                    "Do you want to load it anyway?"
                )
                if not result:
                    return
            
            loaded_count = 0
            for config_entry in config_data.get('configurations', []):
                key = (
                    config_entry['test_value'],
                    config_entry['range_setting'],
                    config_entry['io_type']
                )
                
                if key in entry_widgets:
                    entries = entry_widgets[key]
                    
                    # Clear and set range
                    entries['range'].delete(0, tk.END)
                    entries['range'].insert(0, config_entry.get('range_input', 'N/A'))
                    
                    # Clear and set reference
                    entries['reference'].delete(0, tk.END)
                    entries['reference'].insert(0, config_entry.get('reference', str(key[0])))
                    
                    # Clear and set tolerance
                    entries['tolerance'].delete(0, tk.END)
                    entries['tolerance'].insert(0, config_entry.get('tolerance', '0.015'))
                    
                    loaded_count += 1
            
            status_var.set(f"✓ Loaded {loaded_count} configurations from {Path(file_path).name}")
            root.after(5000, lambda: status_var.set(""))  # Clear after 5 seconds
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to apply configuration:\n{str(e)}")
    
    save_btn = tk.Button(config_row, text="💾 Save Config", command=save_config,
                        font=('Segoe UI', 10), bg='#5B9BD5', fg='white',
                        width=14, height=1, cursor='hand2', relief='raised', bd=2)
    save_btn.pack(side='right', padx=5)
    
    load_btn = tk.Button(config_row, text="📂 Load Config", command=load_config,
                        font=('Segoe UI', 10), bg='#5B9BD5', fg='white',
                        width=14, height=1, cursor='hand2', relief='raised', bd=2)
    load_btn.pack(side='right', padx=5)
    
    # Main action buttons row (Cancel/Submit)
    action_row = tk.Frame(button_frame, bg='#F0F0F0')
    action_row.pack(fill='x', pady=(5, 15))
    
    result = {'cancelled': False}
    
    def on_submit():
        try:
            for (test_value, range_setting, io_type), entries in entry_widgets.items():
                range_str = entries['range'].get().strip()
                final_range = None if range_str.upper() == 'N/A' or range_str == '' else range_str
                
                ref_str = entries['reference'].get().strip().replace(',', '.')
                reference = float(ref_str)
                
                tol_str = entries['tolerance'].get().strip().replace(',', '.')
                tolerance = float(tol_str)
                
                if tolerance < 0:
                    messagebox.showerror("Error", f"Tolerance for {test_value} {unit} ({io_type}) must be positive!")
                    return
                
                user_inputs[(test_value, range_setting, io_type)] = {
                    'range': final_range,
                    'reference': reference,
                    'tolerance': tolerance
                }
            
            result['cancelled'] = False
            root.quit()
            root.destroy()
        except ValueError as e:
            messagebox.showerror("Error", f"Please enter valid numbers!\n{str(e)}")
    
    def on_cancel():
        result['cancelled'] = True
        root.quit()
        root.destroy()
    
    submit_btn = tk.Button(action_row, text="Submit", command=on_submit,
                          font=('Segoe UI', 11, 'bold'), bg='#0070C0', fg='white',
                          width=12, height=2, cursor='hand2', relief='raised', bd=2)
    submit_btn.pack(side='right', padx=20)
    
    cancel_btn = tk.Button(action_row, text="Cancel", command=on_cancel,
                          font=('Segoe UI', 11), bg='#E0E0E0', fg='black',
                          width=12, height=2, cursor='hand2', relief='raised', bd=2)
    cancel_btn.pack(side='right', padx=5)
    
    # Center window
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f'+{x}+{y}')
    
    def on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    
    canvas.bind_all("<MouseWheel>", on_mousewheel)
    
    # Auto-load config if exists in input directory
    if input_dir:
        default_config_path = os.path.join(input_dir, "test_config.json")
        if os.path.exists(default_config_path):
            root.after(100, lambda: apply_config_file(default_config_path))
    
    root.mainloop()
    
    if result['cancelled']:
        return None
    
    return user_inputs


def get_versioned_filename(base_path):
    """
    Generate a versioned filename if the file already exists.
    Returns the next available filename with _v2, _v3, etc.
    """
    if not os.path.exists(base_path):
        return base_path
    
    # Split the path into directory, name, and extension
    directory = os.path.dirname(base_path)
    filename = os.path.basename(base_path)
    name, ext = os.path.splitext(filename)
    
    # Check if filename already has a version suffix
    import re
    version_match = re.match(r'^(.+)_v(\d+)$', name)
    if version_match:
        base_name = version_match.group(1)
        current_version = int(version_match.group(2))
    else:
        base_name = name
        current_version = 1
    
    # Find the next available version
    version = current_version + 1 if current_version > 1 else 2
    while True:
        new_filename = f"{base_name}_v{version}{ext}"
        new_path = os.path.join(directory, new_filename)
        if not os.path.exists(new_path):
            return new_path
        version += 1


def process_files(input_dir='.', user_inputs=None, unit='V', measurement_type_selections=None):
    """
    Process all CSV and TXT files in the input directory and compile results into Excel.
    """
    dir_name = Path(input_dir).name
    if not dir_name:
        dir_name = Path(input_dir).resolve().name
    
    # Generate versioned output filename
    base_output_file = os.path.join(input_dir, f'{dir_name}.xlsx')
    output_file = get_versioned_filename(base_output_file)
    
    results = []
    
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    total_files = len(csv_files) + len(txt_files)
    if total_files == 0:
        print(f"No CSV or TXT files found in {input_dir}")
        return
    
    print(f"Found {len(csv_files)} CSV files and {len(txt_files)} TXT files")
    
    # Process CSV files (output data)
    for csv_file in csv_files:
        value, file_unit, channel, range_setting = parse_filename(csv_file.name)
        
        if value is None or channel is None:
            print(f"Skipping {csv_file.name} - could not parse filename (value or channel missing)")
            continue
        
        try:
            df = pd.read_csv(csv_file)
            
            measurement_col = None
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['voltage', 'vdc', 'resistance', 'ohm', 'current', 'adc', 'measurement']):
                    measurement_col = col
                    break
            
            if measurement_col is None:
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) == 0:
                    print(f"Warning: No numeric columns found in {csv_file.name}")
                    continue
                measurement_col = numeric_cols[-1]
            
            measurements = df[measurement_col].dropna()
            
            if len(measurements) == 0:
                print(f"Warning: No valid measurements in {csv_file.name}")
                continue
            
            result = {
                'Channel': channel,
                'I/O Type': 'Output',
                'Range Setting': range_setting if range_setting else 'N/A',
                f'Test Value [{unit}]': value,
                f'Mean [{unit}]': measurements.mean(),
                f'StdDev [{unit}]': measurements.std(),
                f'Min [{unit}]': measurements.min(),
                f'Max [{unit}]': measurements.max(),
                'Samples': len(measurements),
                '_range_key': range_setting
            }
            
            results.append(result)
            print(f"Processed: {csv_file.name} - CH{channel}, {value}{unit}, Range:{range_setting or 'N/A'}, {len(measurements)} samples (Output)")
            
        except Exception as e:
            print(f"Error processing {csv_file.name}: {str(e)}")
            continue
    
    # Process TXT files (input data)
    for txt_file in txt_files:
        value, file_unit, channel_from_name, range_setting = parse_filename(txt_file.name)
        
        if value is None:
            print(f"Skipping {txt_file.name} - could not parse test value from filename")
            continue
        
        try:
            # Get selected measurement type for this file
            selected_type = None
            if measurement_type_selections and str(txt_file) in measurement_type_selections:
                selected_type = measurement_type_selections[str(txt_file)]
            
            # Parse the text file, passing channel from filename if available
            channel_data = parse_text_file(txt_file, selected_measurement_type=selected_type, 
                                          channel_from_filename=channel_from_name)
            
            if not channel_data:
                print(f"Warning: No valid measurements parsed from {txt_file.name}")
                continue
            
            # Process each channel found in the file
            for channel, measurements in channel_data.items():
                if len(measurements) == 0:
                    continue
                
                measurements = pd.Series(measurements).dropna()
                
                if len(measurements) == 0:
                    continue
                
                result = {
                    'Channel': channel,
                    'I/O Type': 'Input',
                    'Range Setting': range_setting if range_setting else 'N/A',
                    f'Test Value [{unit}]': value,
                    f'Mean [{unit}]': measurements.mean(),
                    f'StdDev [{unit}]': measurements.std(),
                    f'Min [{unit}]': measurements.min(),
                    f'Max [{unit}]': measurements.max(),
                    'Samples': len(measurements),
                    '_range_key': range_setting
                }
                
                results.append(result)
                type_info = f" ({selected_type})" if selected_type else ""
                print(f"Processed: {txt_file.name} - CH{channel}, {value}{unit}, Range:{range_setting or 'N/A'}, {len(measurements)} samples (Input){type_info}")
            
        except Exception as e:
            print(f"Error processing {txt_file.name}: {str(e)}")
            continue
    
    if not results:
        print("No valid results to save")
        return
    
    # Create DataFrame and sort
    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values(['Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]'])
    
    # Add reference value, tolerance, and limits columns
    if user_inputs:
        def get_user_config(row):
            test_val = row[f'Test Value [{unit}]']
            range_key = row['_range_key']
            io_type = row['I/O Type']
            key = (test_val, range_key, io_type)
            return user_inputs.get(key, {})
        
        df_results['_config'] = df_results.apply(get_user_config, axis=1)
        
        df_results[f'Reference Value [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('reference', np.nan) if x else np.nan
        )
        df_results[f'Tolerance [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('tolerance', np.nan) if x else np.nan
        )
        
        df_results['Range Setting'] = df_results.apply(
            lambda row: row['_config'].get('range', row['Range Setting']) 
                        if row['_config'] and row['_config'].get('range') is not None 
                        else row['Range Setting'] if row['Range Setting'] != 'N/A' else 'N/A',
            axis=1
        )
        
        # Calculate limits using reference value
        df_results[f'Lower Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] - df_results[f'Tolerance [{unit}]']
        df_results[f'Upper Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] + df_results[f'Tolerance [{unit}]']
        
        df_results['Mean Check'] = df_results.apply(
            lambda row: 'PASS' if row[f'Lower Limit [{unit}]'] <= row[f'Mean [{unit}]'] <= row[f'Upper Limit [{unit}]'] else 'FAIL',
            axis=1
        )
        
        df_results['Mean±2σ Check'] = df_results.apply(
            lambda row: 'PASS' if (
                row[f'Lower Limit [{unit}]'] <= (row[f'Mean [{unit}]'] - 2*row[f'StdDev [{unit}]']) and
                (row[f'Mean [{unit}]'] + 2*row[f'StdDev [{unit}]']) <= row[f'Upper Limit [{unit}]']
            ) else 'FAIL',
            axis=1
        )
        
        df_results = df_results.drop(columns=['_config', '_range_key'])
        
        column_order = [
            'Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]', 
            f'Reference Value [{unit}]', f'Tolerance [{unit}]', 
            f'Lower Limit [{unit}]', f'Upper Limit [{unit}]',
            f'Mean [{unit}]', f'StdDev [{unit}]', f'Min [{unit}]', f'Max [{unit}]', 
            'Samples', 'Mean Check', 'Mean±2σ Check'
        ]
        df_results = df_results[column_order]
    else:
        df_results = df_results.drop(columns=['_range_key'])
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Test Results', index=False)
        
        worksheet = writer.sheets['Test Results']
        worksheet.auto_filter.ref = worksheet.dimensions
        
        from openpyxl.styles import PatternFill, Font
        
        if user_inputs:
            numeric_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12]
            samples_col = 13
            pass_fail_cols = [14, 15]
        else:
            numeric_cols = [4, 5, 6, 7, 8]
            samples_col = 9
            pass_fail_cols = []
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.number_format = '0.000000'
                elif cell.column == samples_col:
                    cell.number_format = '0'
                elif cell.column in pass_fail_cols:
                    if cell.value == 'PASS':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == 'FAIL':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
        
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    if user_inputs:
        color_assignments = create_tolerance_charts(output_file, df_results, unit)
        
        # Apply channel colors to Test Results sheet
        if color_assignments:
            apply_channel_colors_to_results(output_file, df_results, unit, color_assignments)
        
        # Generate interactive HTML report
        html_file = create_html_report(output_file, df_results, unit)
    else:
        html_file = None
    
    print(f"\n✓ Results saved to {output_file}")
    print(f"  Total entries: {len(df_results)}")
    print(f"  Channels: {sorted(df_results['Channel'].unique())}")
    print(f"  Test values: {sorted(df_results[f'Test Value [{unit}]'].unique())}")
    print(f"  I/O Types: {sorted(df_results['I/O Type'].unique())}")
    
    return output_file, html_file


def apply_channel_colors_to_results(excel_file, df_results, unit, color_assignments):
    """
    Apply channel colors to the Test Results sheet based on the color assignments
    from the tolerance charts.
    """
    from openpyxl.styles import Font, PatternFill
    
    wb = load_workbook(excel_file)
    ws = wb['Test Results']
    
    # Columns A to M (1 to 13)
    color_columns = list(range(1, 14))
    
    # Iterate through data rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        if row_idx >= len(df_results):
            break
        
        # Get the data for this row from the dataframe
        row_data = df_results.iloc[row_idx]
        channel = row_data['Channel']
        io_type = row_data['I/O Type']
        test_value = row_data[f'Test Value [{unit}]']
        range_setting = row_data['Range Setting']
        
        # Look up the color for this combination (range_setting is used as-is, including 'N/A')
        key = (channel, io_type, test_value, range_setting)
        color = color_assignments.get(key)
        
        if color:
            # Apply color to columns A through M
            for cell in row:
                if cell.column in color_columns:
                    # Preserve existing formatting but change font color
                    current_font = cell.font
                    cell.font = Font(
                        name=current_font.name,
                        size=current_font.size,
                        bold=current_font.bold,
                        italic=current_font.italic,
                        color=color
                    )
    
    wb.save(excel_file)
    print("✓ Channel colors applied to Test Results sheet")


def create_tolerance_charts(excel_file, df_results, unit):
    """
    Create tolerance charts showing limits, reference value, mean, and mean±2σ for each 
    test value + range setting combination.
    
    Returns: Dictionary mapping (channel, io_type, test_value, range_setting) to color
    """
    print("\nCreating Tolerance charts...")
    
    wb = load_workbook(excel_file)
    
    if 'Tolerance Charts' in wb.sheetnames:
        del wb['Tolerance Charts']
    chart_sheet = wb.create_sheet('Tolerance Charts')
    
    # Get unique combinations
    unique_combinations = df_results.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    charts_per_row = 2
    chart_width = 26
    chart_height = 36
    
    # Pleasant, muted color palette for channels (same color for mean, +2σ, -2σ)
    channel_colors = [
        '4472C4',  # Muted blue
        'C45B5B',  # Muted red
        '70AD47',  # Muted green
        'ED7D31',  # Muted orange
        '7B7B7B',  # Gray
        '9E5ECE',  # Muted purple
        '43A6A2',  # Muted teal
        'C4A24E',  # Muted gold
        '5B9BC4',  # Steel blue
        'A85B5B',  # Dusty rose
        '5BAF7B',  # Sea green
        'C47B4E',  # Terracotta
        '6B6BAF',  # Muted indigo
        '8B6BAF',  # Muted violet
        '4EAFAF',  # Turquoise
        'AF8B4E',  # Bronze
    ]
    
    # Track color assignments for each channel within each chart context
    # Key: (channel, io_type, test_value, range_setting) -> color
    color_assignments = {}
    
    chart_idx = 0
    
    from openpyxl.styles import Font, Alignment, PatternFill
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df_results[f'Test Value [{unit}]'] == test_value) &
            (df_results['Range Setting'] == range_setting) &
            (df_results['I/O Type'] == io_type)
        )
        test_data = df_results[mask].copy()
        test_data = test_data.sort_values('Channel')
        
        if len(test_data) == 0:
            continue
        
        col_offset = (chart_idx % charts_per_row) * chart_width
        row_offset = (chart_idx // charts_per_row) * chart_height
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        title_row = row_offset + 1
        chart_sheet.cell(title_row, col_offset + 1).value = chart_title
        title_cell = chart_sheet.cell(title_row, col_offset + 1)
        title_cell.font = Font(bold=True, size=11, color='1F4E78')
        title_cell.alignment = Alignment(horizontal='left')
        
        channels = test_data['Channel'].tolist()
        num_channels = len(channels)
        lower_limits = test_data[f'Lower Limit [{unit}]'].tolist()
        upper_limits = test_data[f'Upper Limit [{unit}]'].tolist()
        reference_values = test_data[f'Reference Value [{unit}]'].tolist()
        means = test_data[f'Mean [{unit}]'].tolist()
        lower_2sigma = (test_data[f'Mean [{unit}]'] - 2*test_data[f'StdDev [{unit}]']).tolist()
        upper_2sigma = (test_data[f'Mean [{unit}]'] + 2*test_data[f'StdDev [{unit}]']).tolist()
        mean_checks = test_data['Mean Check'].tolist()
        mean_2sigma_checks = test_data['Mean±2σ Check'].tolist()
        
        # Get limit values (same for all channels in this chart)
        ref_val = reference_values[0]
        ll_val = lower_limits[0]
        ul_val = upper_limits[0]
        
        data_start_row = row_offset + 3
        data_start_col = col_offset + 1
        
        # Write headers
        headers = ['Channel', 'Lower Limit', 'Reference', 'Upper Limit', 'Mean', 'Mean-2σ', 'Mean+2σ', 'Mean Check', 'Mean±2σ Check']
        for h_idx, header in enumerate(headers):
            chart_sheet.cell(data_start_row, data_start_col + h_idx).value = header
            chart_sheet.cell(data_start_row, data_start_col + h_idx).font = Font(bold=True, size=9)
        
        # Write data with color-coded fonts
        for i, channel in enumerate(channels):
            row = data_start_row + i + 1
            color = channel_colors[i % len(channel_colors)]
            
            # Store color assignment
            color_assignments[(channel, io_type, test_value, range_setting)] = color
            
            # Apply color to all cells in this row
            cell_font = Font(size=9, color=color)
            
            chart_sheet.cell(row, data_start_col, channel).font = cell_font
            chart_sheet.cell(row, data_start_col + 1, lower_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 1).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 2, reference_values[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 2).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 3, upper_limits[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 3).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 4, means[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 4).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 5, lower_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 5).number_format = '0.000000'
            chart_sheet.cell(row, data_start_col + 6, upper_2sigma[i]).font = cell_font
            chart_sheet.cell(row, data_start_col + 6).number_format = '0.000000'
            
            # Mean Check column with PASS/FAIL formatting
            mean_check_cell = chart_sheet.cell(row, data_start_col + 7, mean_checks[i])
            if mean_checks[i] == 'PASS':
                mean_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # Mean±2σ Check column with PASS/FAIL formatting
            mean_2sigma_check_cell = chart_sheet.cell(row, data_start_col + 8, mean_2sigma_checks[i])
            if mean_2sigma_checks[i] == 'PASS':
                mean_2sigma_check_cell.font = Font(size=9, color='006100', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                mean_2sigma_check_cell.font = Font(size=9, color='9C0006', bold=True)
                mean_2sigma_check_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Create scatter chart
        chart = ScatterChart()
        chart.title = chart_title
        chart.style = 10
        
        # Set axis titles (non-bold is default for axis titles)
        chart.x_axis.title = "Channel"
        chart.y_axis.title = f"Measurement [{unit}]"
        
        # Chart size - make it larger for better visibility
        chart.height = 14
        chart.width = 18
        
        # Remove legend
        chart.legend = None
        
        # Remove gridlines
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None
        
        # Calculate Y-axis limits with padding
        all_values = lower_limits + upper_limits + reference_values + means + lower_2sigma + upper_2sigma
        y_min = min(all_values)
        y_max = max(all_values)
        y_range = y_max - y_min if y_max != y_min else abs(y_max) * 0.1 or 0.1
        y_padding = y_range * 0.20  # 20% padding
        
        chart.y_axis.scaling.min = y_min - y_padding
        chart.y_axis.scaling.max = y_max + y_padding
        
        # Set X-axis to properly scale for the number of channels
        x_min_val = min(channels)
        x_max_val = max(channels)
        x_padding_left = max(0.8, (x_max_val - x_min_val) * 0.1)
        x_padding_right = max(0.8, (x_max_val - x_min_val) * 0.1)
        chart.x_axis.scaling.min = x_min_val - x_padding_left
        chart.x_axis.scaling.max = x_max_val + x_padding_right
        chart.x_axis.majorUnit = 1  # Show each channel number
        
        # Reference for x values (channel numbers)
        xvalues = Reference(chart_sheet, min_col=data_start_col, min_row=data_start_row+1, max_row=data_start_row+num_channels)
        
        # Lower limit line (dark red, dashed)
        lower_series = Series(Reference(chart_sheet, min_col=data_start_col+1, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Lower Limit")
        lower_series.marker = Marker('none')
        lower_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(lower_series)
        
        # Reference value line (dark green, solid)
        ref_series = Series(Reference(chart_sheet, min_col=data_start_col+2, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                           xvalues, title="Reference")
        ref_series.marker = Marker('none')
        ref_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="2E7D32", w=12700))
        chart.series.append(ref_series)
        
        # Upper limit line (dark red, dashed)
        upper_series = Series(Reference(chart_sheet, min_col=data_start_col+3, min_row=data_start_row+1, max_row=data_start_row+num_channels), 
                            xvalues, title="Upper Limit")
        upper_series.marker = Marker('none')
        upper_series.graphicalProperties = GraphicalProperties(ln=LineProperties(solidFill="8B0000", w=12700, prstDash='dash'))
        chart.series.append(upper_series)
        
        # Add series for each channel with consistent colors
        for i, channel in enumerate(channels):
            color = channel_colors[i % len(channel_colors)]
            channel_row = data_start_row + 1 + i
            
            # Mean (diamond marker - smaller size)
            mean_series = Series(Reference(chart_sheet, min_col=data_start_col+4, min_row=channel_row, max_row=channel_row), 
                               Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                               title=f"CH{channel} Mean")
            mean_series.marker = Marker('diamond', size=6)
            mean_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color))
            mean_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(mean_series)
            
            # Mean-2σ (horizontal line marker - thinner)
            lower_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+5, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} -2σ")
            lower_2s_series.marker = Marker('dash', size=8)
            lower_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            lower_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(lower_2s_series)
            
            # Mean+2σ (horizontal line marker - thinner)
            upper_2s_series = Series(Reference(chart_sheet, min_col=data_start_col+6, min_row=channel_row, max_row=channel_row), 
                                   Reference(chart_sheet, min_col=data_start_col, min_row=channel_row, max_row=channel_row), 
                                   title=f"CH{channel} +2σ")
            upper_2s_series.marker = Marker('dash', size=8)
            upper_2s_series.marker.graphicalProperties = GraphicalProperties(solidFill=color, ln=LineProperties(solidFill=color, w=12700))
            upper_2s_series.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
            chart.series.append(upper_2s_series)
        
        # Position chart (moved further right to accommodate check columns)
        chart_cell = chart_sheet.cell(row_offset + 4, col_offset + 12)
        chart.anchor = chart_cell.coordinate
        chart_sheet.add_chart(chart)
        
        print(f"  Created chart for {chart_title}")
        chart_idx += 1
    
    wb.save(excel_file)
    print("✓ Tolerance charts added to workbook")
    
    return color_assignments


def create_html_report(output_file, df_results, unit):
    """
    Create an interactive HTML report using Plotly with tolerance charts and data tables.
    """
    if not PLOTLY_AVAILABLE:
        print("Warning: Plotly not available. Skipping HTML report generation.")
        print("  Install with: pip install plotly")
        return None
    
    print("\nCreating interactive HTML report...")
    
    # Generate HTML filename
    html_file = output_file.replace('.xlsx', '_report.html')
    
    # Pleasant, muted color palette for channels (matching Excel charts)
    channel_colors = [
        '#4472C4',  # Muted blue
        '#C45B5B',  # Muted red
        '#70AD47',  # Muted green
        '#ED7D31',  # Muted orange
        '#7B7B7B',  # Gray
        '#9E5ECE',  # Muted purple
        '#43A6A2',  # Muted teal
        '#C4A24E',  # Muted gold
        '#5B9BC4',  # Steel blue
        '#A85B5B',  # Dusty rose
        '#5BAF7B',  # Sea green
        '#C47B4E',  # Terracotta
        '#6B6BAF',  # Muted indigo
        '#8B6BAF',  # Muted violet
        '#4EAFAF',  # Turquoise
        '#AF8B4E',  # Bronze
    ]
    
    # Get unique combinations for charts
    unique_combinations = df_results.groupby([f'Test Value [{unit}]', 'Range Setting', 'I/O Type']).size().reset_index()
    unique_combinations = unique_combinations.sort_values([f'Test Value [{unit}]', 'Range Setting', 'I/O Type'])
    
    num_charts = len(unique_combinations)
    
    # Create figures list
    figures_html = []
    
    for _, combo_row in unique_combinations.iterrows():
        test_value = combo_row[f'Test Value [{unit}]']
        range_setting = combo_row['Range Setting']
        io_type = combo_row['I/O Type']
        
        mask = (
            (df_results[f'Test Value [{unit}]'] == test_value) &
            (df_results['Range Setting'] == range_setting) &
            (df_results['I/O Type'] == io_type)
        )
        test_data = df_results[mask].copy()
        test_data = test_data.sort_values('Channel')
        
        if len(test_data) == 0:
            continue
        
        channels = test_data['Channel'].tolist()
        lower_limits = test_data[f'Lower Limit [{unit}]'].tolist()
        upper_limits = test_data[f'Upper Limit [{unit}]'].tolist()
        reference_values = test_data[f'Reference Value [{unit}]'].tolist()
        means = test_data[f'Mean [{unit}]'].tolist()
        stddevs = test_data[f'StdDev [{unit}]'].tolist()
        lower_2sigma = [m - 2*s for m, s in zip(means, stddevs)]
        upper_2sigma = [m + 2*s for m, s in zip(means, stddevs)]
        mean_checks = test_data['Mean Check'].tolist()
        mean_2sigma_checks = test_data['Mean±2σ Check'].tolist()
        
        # Get limit values (same for all channels)
        ref_val = reference_values[0]
        ll_val = lower_limits[0]
        ul_val = upper_limits[0]
        
        range_display = f", Range: {range_setting}" if range_setting != 'N/A' else ""
        chart_title = f"Test: {test_value} {unit}{range_display} ({io_type})"
        
        # Create figure
        fig = go.Figure()
        
        # Add limit lines (horizontal lines across all channels)
        x_range = [min(channels) - 0.5, max(channels) + 0.5]
        
        # Lower limit line (dashed red)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ll_val, ll_val],
            mode='lines',
            name=f'Lower Limit ({ll_val:.6f})',
            line=dict(color='#8B0000', width=2, dash='dash'),
            hoverinfo='name+y'
        ))
        
        # Reference line (solid green)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ref_val, ref_val],
            mode='lines',
            name=f'Reference ({ref_val:.6f})',
            line=dict(color='#2E7D32', width=2),
            hoverinfo='name+y'
        ))
        
        # Upper limit line (dashed red)
        fig.add_trace(go.Scatter(
            x=x_range,
            y=[ul_val, ul_val],
            mode='lines',
            name=f'Upper Limit ({ul_val:.6f})',
            line=dict(color='#8B0000', width=2, dash='dash'),
            hoverinfo='name+y'
        ))
        
        # Add data points for each channel
        for i, channel in enumerate(channels):
            color = channel_colors[i % len(channel_colors)]
            
            # Mean point (diamond)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[means[i]],
                mode='markers',
                name=f'CH{channel} Mean',
                marker=dict(
                    symbol='diamond',
                    size=12,
                    color=color,
                    line=dict(color=color, width=1)
                ),
                hovertemplate=f'CH{channel}<br>Mean: %{{y:.6f}}<br>Check: {mean_checks[i]}<extra></extra>'
            ))
            
            # Mean-2σ point (line marker)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[lower_2sigma[i]],
                mode='markers',
                name=f'CH{channel} -2σ',
                marker=dict(
                    symbol='line-ew',
                    size=10,
                    color=color,
                    line=dict(color=color, width=3)
                ),
                hovertemplate=f'CH{channel}<br>Mean-2σ: %{{y:.6f}}<extra></extra>',
                showlegend=False
            ))
            
            # Mean+2σ point (line marker)
            fig.add_trace(go.Scatter(
                x=[channel],
                y=[upper_2sigma[i]],
                mode='markers',
                name=f'CH{channel} +2σ',
                marker=dict(
                    symbol='line-ew',
                    size=10,
                    color=color,
                    line=dict(color=color, width=3)
                ),
                hovertemplate=f'CH{channel}<br>Mean+2σ: %{{y:.6f}}<br>±2σ Check: {mean_2sigma_checks[i]}<extra></extra>',
                showlegend=False
            ))
            
            # Add vertical line connecting -2σ to +2σ
            fig.add_trace(go.Scatter(
                x=[channel, channel],
                y=[lower_2sigma[i], upper_2sigma[i]],
                mode='lines',
                line=dict(color=color, width=1),
                showlegend=False,
                hoverinfo='skip'
            ))
        
        # Calculate Y-axis range
        all_values = lower_limits + upper_limits + reference_values + means + lower_2sigma + upper_2sigma
        y_min = min(all_values)
        y_max = max(all_values)
        y_range_val = y_max - y_min if y_max != y_min else abs(y_max) * 0.1 or 0.1
        y_padding = y_range_val * 0.20
        
        # Update layout - title moved outside chart, legends hidden by default
        fig.update_layout(
            title=None,  # Title will be added as external HTML element
            xaxis=dict(
                title='Channel',
                tickmode='linear',
                tick0=min(channels),
                dtick=1,
                autorange=True  # Enable autoscale
            ),
            yaxis=dict(
                title=f'Measurement [{unit}]',
                autorange=True  # Enable autoscale
            ),
            legend=dict(
                orientation='h',
                yanchor='bottom',
                y=1.0,
                xanchor='center',
                x=0.5,
                font=dict(size=10),
                bgcolor='rgba(255,255,255,0.9)',
                bordercolor='#e9ecef',
                borderwidth=1
            ),
            showlegend=False,  # Legends hidden by default
            hovermode='closest',
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin=dict(l=60, r=20, t=30, b=50),
            autosize=True
        )
        
        # Add gridlines
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='#E0E0E0')
        
        # Store chart with its title and io_type
        figures_html.append({
            'title': chart_title,
            'io_type': io_type,
            'html': fig.to_html(full_html=False, include_plotlyjs=False, config={'responsive': True})
        })
    
    # Create summary statistics table
    summary_pass = (df_results['Mean Check'] == 'PASS').sum()
    summary_fail = (df_results['Mean Check'] == 'FAIL').sum()
    summary_2s_pass = (df_results['Mean±2σ Check'] == 'PASS').sum()
    summary_2s_fail = (df_results['Mean±2σ Check'] == 'FAIL').sum()
    
    # Get unique values for filters
    unique_channels = sorted(df_results['Channel'].unique())
    unique_ranges = sorted(df_results['Range Setting'].unique())
    unique_test_values = sorted(df_results[f'Test Value [{unit}]'].unique())
    unique_io_types = sorted(df_results['I/O Type'].unique())
    
    # Build HTML document
    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Measurement Report - {Path(output_file).stem}</title>
    <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
    <style>
        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background-color: #f5f7fa;
            color: #333;
            line-height: 1.6;
        }}
        .header {{
            background: linear-gradient(135deg, #5C2D91 0%, #9B59B6 50%, #E8E0F0 100%);
            color: white;
            padding: 30px 40px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
            text-shadow: 1px 1px 2px rgba(0,0,0,0.2);
        }}
        .header p {{
            opacity: 0.9;
            font-size: 14px;
        }}
        .container {{
            width: 100%;
            padding: 20px 30px;
        }}
        .summary-cards {{
            display: grid;
            grid-template-columns: repeat(5, 1fr);
            gap: 20px;
            margin: 20px 0;
        }}
        @media (max-width: 1200px) {{
            .summary-cards {{
                grid-template-columns: repeat(3, 1fr);
            }}
        }}
        @media (max-width: 768px) {{
            .summary-cards {{
                grid-template-columns: repeat(2, 1fr);
            }}
        }}
        .card {{
            background: white;
            border-radius: 8px;
            padding: 12px 15px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            text-align: center;
        }}
        .card h3 {{
            font-size: 11px;
            color: #666;
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .card .value {{
            font-size: 24px;
            font-weight: bold;
        }}
        .card.pass .value {{
            color: #2E7D32;
        }}
        .card.fail .value {{
            color: #C00000;
        }}
        .card.neutral .value {{
            color: #1F4E78;
        }}
        .section {{
            background: white;
            border-radius: 10px;
            margin: 20px 0;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            overflow: hidden;
        }}
        .section-header {{
            background: #f8f9fa;
            padding: 15px 20px;
            border-bottom: 1px solid #e9ecef;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        .section-header:hover {{
            background: #e9ecef;
        }}
        .section-header h2 {{
            font-size: 18px;
            color: #1F4E78;
        }}
        .section-header .toggle {{
            font-size: 20px;
            color: #666;
        }}
        .section-content {{
            padding: 20px;
        }}
        .chart-grid {{
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 25px;
        }}
        @media (max-width: 1400px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
        }}
        .chart-container {{
            background: #fafafa;
            border-radius: 8px;
            padding: 15px;
            border: 1px solid #e9ecef;
            min-height: 400px;
        }}
        .chart-title {{
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 10px;
            padding: 8px 12px;
            border-radius: 0 4px 4px 0;
        }}
        .chart-title.input {{
            color: #2E7D32;
            background: linear-gradient(90deg, #e8f5e9 0%, transparent 100%);
            border-left: 4px solid #2E7D32;
        }}
        .chart-title.output {{
            color: #1F4E78;
            background: linear-gradient(90deg, #e0f0ff 0%, transparent 100%);
            border-left: 4px solid #1F4E78;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 380px;
        }}
        .chart-wrapper > div {{
            width: 100% !important;
            height: 100% !important;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }}
        th, td {{
            padding: 10px 12px;
            text-align: left;
            border-bottom: 1px solid #e9ecef;
        }}
        th {{
            background: #f8f9fa;
            font-weight: 600;
            color: #1F4E78;
            position: sticky;
            top: 0;
            white-space: nowrap;
        }}
        tr:hover {{
            background: #f8f9fa;
        }}
        .pass {{
            background-color: #C6EFCE;
            color: #006100;
            font-weight: bold;
            text-align: center;
            border-radius: 4px;
        }}
        .fail {{
            background-color: #FFC7CE;
            color: #9C0006;
            font-weight: bold;
            text-align: center;
            border-radius: 4px;
        }}
        .table-wrapper {{
            max-height: 500px;
            overflow-y: auto;
            border: 1px solid #e9ecef;
            border-radius: 8px;
        }}
        .chart-controls {{
            display: flex;
            align-items: center;
            gap: 20px;
            margin-bottom: 15px;
            flex-wrap: wrap;
        }}
        .toggle-legend-btn {{
            padding: 8px 16px;
            background: #1F4E78;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            transition: background 0.2s;
            display: flex;
            align-items: center;
            gap: 6px;
        }}
        .toggle-legend-btn:hover {{
            background: #2E7D32;
        }}
        .toggle-legend-btn.legends-hidden {{
            background: #666;
        }}
        .filter-bar {{
            display: flex;
            gap: 10px;
            margin-bottom: 15px;
            flex-wrap: wrap;
            align-items: center;
        }}
        .filter-group {{
            display: flex;
            align-items: center;
            gap: 6px;
            background: #f8f9fa;
            padding: 6px 10px;
            border-radius: 6px;
        }}
        .filter-bar label {{
            font-weight: 500;
            color: #666;
            font-size: 13px;
            white-space: nowrap;
        }}
        .filter-bar select, .filter-bar input {{
            padding: 6px 10px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 13px;
            min-width: 100px;
        }}
        .filter-bar input {{
            min-width: 150px;
        }}
        .filter-bar select:focus, .filter-bar input:focus {{
            outline: none;
            border-color: #1F4E78;
        }}
        .clear-filters-btn {{
            padding: 6px 12px;
            background: #e9ecef;
            color: #666;
            border: 1px solid #ddd;
            border-radius: 4px;
            cursor: pointer;
            font-size: 13px;
            transition: all 0.2s;
        }}
        .clear-filters-btn:hover {{
            background: #ddd;
            color: #333;
        }}
        .filter-count {{
            font-size: 12px;
            color: #666;
            padding: 4px 8px;
            background: #e9ecef;
            border-radius: 4px;
        }}
        .legend-info {{
            display: flex;
            gap: 20px;
            flex-wrap: wrap;
            padding: 10px 15px;
            background: #f8f9fa;
            border-radius: 8px;
            font-size: 13px;
        }}
        .legend-item {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .legend-line {{
            width: 30px;
            height: 3px;
        }}
        .legend-marker {{
            width: 12px;
            height: 12px;
        }}
        .collapsible {{
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.5s ease-out;
        }}
        .collapsible.active {{
            max-height: none;
            overflow: visible;
        }}
        @media (max-width: 768px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
            .header {{
                padding: 20px;
            }}
            .header h1 {{
                font-size: 22px;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 Measurement Analysis Report</h1>
        <p>Generated from: {Path(output_file).stem} | Unit: {unit} | Total Measurements: {len(df_results)}</p>
    </div>
    
    <div class="container">
        <!-- Summary Cards -->
        <div class="summary-cards">
            <div class="card neutral">
                <h3>Total Tests</h3>
                <div class="value">{len(df_results)}</div>
            </div>
            <div class="card pass">
                <h3>Mean Check Pass</h3>
                <div class="value">{summary_pass}</div>
            </div>
            <div class="card fail">
                <h3>Mean Check Fail</h3>
                <div class="value">{summary_fail}</div>
            </div>
            <div class="card pass">
                <h3>±2σ Check Pass</h3>
                <div class="value">{summary_2s_pass}</div>
            </div>
            <div class="card fail">
                <h3>±2σ Check Fail</h3>
                <div class="value">{summary_2s_fail}</div>
            </div>
        </div>
        
        <!-- Charts Section -->
        <div class="section">
            <div class="section-header" onclick="toggleSection('charts-section')">
                <h2>📈 Tolerance Charts</h2>
                <span class="toggle" id="charts-section-toggle">▼</span>
            </div>
            <div class="section-content collapsible active" id="charts-section">
                <div class="chart-controls">
                    <button class="toggle-legend-btn legends-hidden" onclick="toggleAllLegends()">
                        <span id="legend-btn-icon">👁️‍🗨️</span> Toggle Legends
                    </button>
                    <div class="legend-info">
                        <div class="legend-item">
                            <div class="legend-line" style="background: #8B0000; border-style: dashed;"></div>
                            <span>Upper/Lower Limits</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-line" style="background: #2E7D32;"></div>
                            <span>Reference Value</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-marker" style="background: #4472C4; transform: rotate(45deg);"></div>
                            <span>Mean</span>
                        </div>
                        <div class="legend-item">
                            <div class="legend-line" style="background: #4472C4; height: 2px;"></div>
                            <span>Mean ± 2σ</span>
                        </div>
                    </div>
                </div>
                <div class="chart-grid">
'''
    
    # Add each chart with external title (color based on I/O type)
    for i, chart_data in enumerate(figures_html):
        io_class = 'input' if chart_data['io_type'] == 'Input' else 'output'
        html_content += f'''
                    <div class="chart-container">
                        <div class="chart-title {io_class}">{chart_data['title']}</div>
                        <div class="chart-wrapper">
                            {chart_data['html']}
                        </div>
                    </div>
'''
    
    html_content += '''
                </div>
            </div>
        </div>
        
        <!-- Data Table Section -->
        <div class="section">
            <div class="section-header" onclick="toggleSection('data-section')">
                <h2>📋 Detailed Results</h2>
                <span class="toggle" id="data-section-toggle">▼</span>
            </div>
            <div class="section-content collapsible active" id="data-section">
                <div class="filter-bar">
                    <div class="filter-group">
                        <label>Channel:</label>
                        <select id="channel-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add channel options
    for ch in unique_channels:
        html_content += f'                            <option value="{ch}">{ch}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>I/O Type:</label>
                        <select id="io-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add I/O type options
    for io in unique_io_types:
        html_content += f'                            <option value="{io}">{io}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Range:</label>
                        <select id="range-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add range options
    for rng in unique_ranges:
        html_content += f'                            <option value="{rng}">{rng}</option>\n'
    
    html_content += f'''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Test Value:</label>
                        <select id="testvalue-filter" onchange="filterTable()">
                            <option value="all">All</option>
'''
    
    # Add test value options
    for tv in unique_test_values:
        html_content += f'                            <option value="{tv}">{tv} {unit}</option>\n'
    
    html_content += '''                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Status:</label>
                        <select id="status-filter" onchange="filterTable()">
                            <option value="all">All</option>
                            <option value="pass">Pass Only</option>
                            <option value="fail">Fail Only</option>
                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Search:</label>
                        <input type="text" id="search-input" placeholder="Search..." onkeyup="filterTable()">
                    </div>
                    <button class="clear-filters-btn" onclick="clearFilters()">Clear All</button>
                    <span class="filter-count" id="filter-count"></span>
                </div>
                <div class="table-wrapper">
                    <table id="results-table">
                        <thead>
                            <tr>
                                <th>Channel</th>
                                <th>I/O Type</th>
                                <th>Range</th>
'''
    
    html_content += f'''
                                <th>Test Value [{unit}]</th>
                                <th>Reference [{unit}]</th>
                                <th>Tolerance [{unit}]</th>
                                <th>Lower Limit [{unit}]</th>
                                <th>Upper Limit [{unit}]</th>
                                <th>Mean [{unit}]</th>
                                <th>StdDev [{unit}]</th>
                                <th>Min [{unit}]</th>
                                <th>Max [{unit}]</th>
                                <th>Samples</th>
                                <th>Mean Check</th>
                                <th>Mean±2σ Check</th>
                            </tr>
                        </thead>
                        <tbody>
'''
    
    # Add table rows
    for _, row in df_results.iterrows():
        mean_class = 'pass' if row['Mean Check'] == 'PASS' else 'fail'
        sigma_class = 'pass' if row['Mean±2σ Check'] == 'PASS' else 'fail'
        
        html_content += f'''
                            <tr>
                                <td>{row['Channel']}</td>
                                <td>{row['I/O Type']}</td>
                                <td>{row['Range Setting']}</td>
                                <td>{row[f'Test Value [{unit}]']:.6f}</td>
                                <td>{row[f'Reference Value [{unit}]']:.6f}</td>
                                <td>{row[f'Tolerance [{unit}]']:.6f}</td>
                                <td>{row[f'Lower Limit [{unit}]']:.6f}</td>
                                <td>{row[f'Upper Limit [{unit}]']:.6f}</td>
                                <td>{row[f'Mean [{unit}]']:.6f}</td>
                                <td>{row[f'StdDev [{unit}]']:.6f}</td>
                                <td>{row[f'Min [{unit}]']:.6f}</td>
                                <td>{row[f'Max [{unit}]']:.6f}</td>
                                <td>{row['Samples']}</td>
                                <td class="{mean_class}">{row['Mean Check']}</td>
                                <td class="{sigma_class}">{row['Mean±2σ Check']}</td>
                            </tr>
'''
    
    html_content += '''
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </div>
    
    <script>
        let legendsVisible = false;
        
        function toggleSection(sectionId) {
            const section = document.getElementById(sectionId);
            const toggle = document.getElementById(sectionId + '-toggle');
            section.classList.toggle('active');
            toggle.textContent = section.classList.contains('active') ? '▼' : '▶';
        }
        
        function toggleAllLegends() {
            legendsVisible = !legendsVisible;
            const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
            const btn = document.querySelector('.toggle-legend-btn');
            const icon = document.getElementById('legend-btn-icon');
            
            charts.forEach(function(chart) {
                Plotly.relayout(chart, { showlegend: legendsVisible });
            });
            
            if (legendsVisible) {
                btn.classList.remove('legends-hidden');
                icon.textContent = '👁️';
            } else {
                btn.classList.add('legends-hidden');
                icon.textContent = '👁️‍🗨️';
            }
        }
        
        function filterTable() {
            const channelFilter = document.getElementById('channel-filter').value;
            const ioFilter = document.getElementById('io-filter').value;
            const rangeFilter = document.getElementById('range-filter').value;
            const testValueFilter = document.getElementById('testvalue-filter').value;
            const statusFilter = document.getElementById('status-filter').value;
            const searchInput = document.getElementById('search-input').value.toLowerCase();
            const table = document.getElementById('results-table');
            const rows = table.getElementsByTagName('tr');
            
            let visibleCount = 0;
            let totalCount = rows.length - 1; // Exclude header
            
            for (let i = 1; i < rows.length; i++) {
                const row = rows[i];
                const cells = row.getElementsByTagName('td');
                
                let showRow = true;
                
                // Channel filter (column 0)
                if (channelFilter !== 'all' && showRow) {
                    const channel = cells[0].textContent;
                    if (channel !== channelFilter) showRow = false;
                }
                
                // I/O Type filter (column 1)
                if (ioFilter !== 'all' && showRow) {
                    const ioType = cells[1].textContent;
                    if (ioType !== ioFilter) showRow = false;
                }
                
                // Range filter (column 2)
                if (rangeFilter !== 'all' && showRow) {
                    const range = cells[2].textContent;
                    if (range !== rangeFilter) showRow = false;
                }
                
                // Test Value filter (column 3)
                if (testValueFilter !== 'all' && showRow) {
                    const testValue = parseFloat(cells[3].textContent);
                    const filterValue = parseFloat(testValueFilter);
                    if (Math.abs(testValue - filterValue) > 0.000001) showRow = false;
                }
                
                // Status filter (columns 13 and 14)
                if (statusFilter !== 'all' && showRow) {
                    const meanCheck = cells[13].textContent;
                    const sigmaCheck = cells[14].textContent;
                    if (statusFilter === 'pass') {
                        if (meanCheck !== 'PASS' || sigmaCheck !== 'PASS') showRow = false;
                    } else if (statusFilter === 'fail') {
                        if (meanCheck !== 'FAIL' && sigmaCheck !== 'FAIL') showRow = false;
                    }
                }
                
                // Search filter
                if (searchInput && showRow) {
                    let found = false;
                    for (let j = 0; j < cells.length; j++) {
                        if (cells[j].textContent.toLowerCase().includes(searchInput)) {
                            found = true;
                            break;
                        }
                    }
                    if (!found) showRow = false;
                }
                
                row.style.display = showRow ? '' : 'none';
                if (showRow) visibleCount++;
            }
            
            // Update filter count
            const countEl = document.getElementById('filter-count');
            if (visibleCount === totalCount) {
                countEl.textContent = `Showing all ${totalCount} rows`;
            } else {
                countEl.textContent = `Showing ${visibleCount} of ${totalCount} rows`;
            }
        }
        
        function clearFilters() {
            document.getElementById('channel-filter').value = 'all';
            document.getElementById('io-filter').value = 'all';
            document.getElementById('range-filter').value = 'all';
            document.getElementById('testvalue-filter').value = 'all';
            document.getElementById('status-filter').value = 'all';
            document.getElementById('search-input').value = '';
            filterTable();
        }
        
        // Resize all Plotly charts when window resizes
        window.addEventListener('resize', function() {
            const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
            charts.forEach(function(chart) {
                Plotly.Plots.resize(chart);
            });
        });
        
        // Initial setup after page load
        window.addEventListener('load', function() {
            setTimeout(function() {
                const charts = document.querySelectorAll('.chart-wrapper .plotly-graph-div');
                charts.forEach(function(chart) {
                    Plotly.Plots.resize(chart);
                });
                // Initialize filter count
                filterTable();
            }, 100);
        });
    </script>
</body>
</html>
'''
    
    # Write HTML file
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ Interactive HTML report saved to {html_file}")
    return html_file


if __name__ == "__main__":
    print("=" * 70)
    print("Measurement Data Compiler with Tolerance Charts")
    print("Supports CSV (Output) and TXT (Input) files with Range Settings")
    print("=" * 70)
    print("\nPlease select the directory containing data files...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    input_dir = filedialog.askdirectory(
        title="Select Directory with CSV/TXT Files",
        initialdir=os.getcwd()
    )
    
    root.destroy()
    
    if not input_dir:
        print("\nNo directory selected. Exiting...")
        exit(0)
    
    print(f"\nSelected directory: {input_dir}")
    
    # Determine unit from files
    unit = get_unit_from_files(input_dir)
    print(f"Detected unit: {unit}")
    
    # Scan files
    print("\nScanning data files...")
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    all_files = csv_files + txt_files
    
    if not all_files:
        print(f"Error: No CSV or TXT files found in {input_dir}")
        exit(1)
    
    # Scan text files for measurement types
    print("Analyzing text file structures...")
    file_measurement_types = {}
    for txt_file in txt_files:
        types = scan_text_file_for_measurement_types(txt_file)
        if types:
            file_measurement_types[str(txt_file)] = types
            if len(types) > 1:
                print(f"  {txt_file.name}: Found multiple measurement types: {', '.join(sorted(types))}")
            else:
                print(f"  {txt_file.name}: Found measurement type: {list(types)[0]}")
    
    # If any file has multiple measurement types, ask user to select
    measurement_type_selections = None
    if file_measurement_types:
        measurement_type_selections = select_measurement_type(file_measurement_types)
        if measurement_type_selections is None:
            print("\nMeasurement type selection cancelled. Exiting...")
            exit(0)
    
    # Extract unique (test_value, range_setting, io_type) tuples from filenames
    test_value_range_io_tuples = set()
    
    # CSV files are Output devices
    for csv_file in csv_files:
        value, _, channel, range_setting = parse_filename(csv_file.name)
        if value is not None and channel is not None:
            test_value_range_io_tuples.add((value, range_setting, 'Output'))
    
    # TXT files are Input devices
    for txt_file in txt_files:
        value, _, _, range_setting = parse_filename(txt_file.name)
        if value is not None:
            test_value_range_io_tuples.add((value, range_setting, 'Input'))
    
    if not test_value_range_io_tuples:
        print("Error: Could not parse any valid test values from filenames")
        exit(1)
    
    print(f"\nFound {len(all_files)} data files ({len(csv_files)} CSV, {len(txt_files)} TXT)")
    print(f"Found {len(test_value_range_io_tuples)} unique test value/range/IO-type combinations")
    
    # Get user inputs
    user_inputs = get_user_inputs(test_value_range_io_tuples, unit, input_dir=input_dir)
    
    if user_inputs is None:
        print("\nConfiguration input cancelled. Exiting...")
        exit(0)
    
    print("\nProcessing files...")
    
    output_file, html_file = process_files(
        input_dir=input_dir, 
        user_inputs=user_inputs, 
        unit=unit,
        measurement_type_selections=measurement_type_selections
    )
    
    print("\n" + "=" * 70)
    print(f"Output file: {output_file}")
    if html_file:
        print(f"HTML report: {html_file}")
        # Automatically open the HTML report in the default browser
        import webbrowser
        webbrowser.open('file://' + os.path.abspath(html_file))
        print("\n✓ HTML report opened in browser")