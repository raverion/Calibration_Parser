"""
Flask Web Application for Measurement Data Analysis
Converts the tkinter-based GUI to a web interface while preserving all backend functionality.
"""

import os
import json
import uuid
import shutil
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font

# Import from local modules (same as original)
from parsers import (
    parse_filename,
    get_unit_from_files,
    scan_text_file_for_measurement_types,
    parse_text_file
)
from excel_charts import create_tolerance_charts, apply_channel_colors_to_results
from html_report import create_html_report
from utils import get_versioned_filename

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Configuration
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
OUTPUT_FOLDER = Path(__file__).parent / 'outputs'
UPLOAD_FOLDER.mkdir(exist_ok=True)
OUTPUT_FOLDER.mkdir(exist_ok=True)

app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)
app.config['OUTPUT_FOLDER'] = str(OUTPUT_FOLDER)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max upload


def get_session_folder():
    """Get or create a unique folder for this session's uploads."""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    
    session_folder = UPLOAD_FOLDER / session['session_id']
    session_folder.mkdir(exist_ok=True)
    return session_folder


def get_output_folder():
    """Get or create a unique folder for this session's outputs."""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
    
    output_folder = OUTPUT_FOLDER / session['session_id']
    output_folder.mkdir(exist_ok=True)
    return output_folder


@app.route('/')
def index():
    """Landing page with mode selection."""
    return render_template('index.html')


@app.route('/equipment-report')
def equipment_report():
    """Equipment-specific report page - upload files."""
    # Clear previous session data
    session.pop('files_info', None)
    session.pop('measurement_types', None)
    session.pop('test_configs', None)
    return render_template('upload.html')


@app.route('/comparison-report')
def comparison_report():
    """Cross-equipment comparison report - disabled for now."""
    return render_template('coming_soon.html')


@app.route('/api/upload', methods=['POST'])
def upload_files():
    """Handle file uploads."""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files selected'}), 400
    
    session_folder = get_session_folder()
    
    # Clear previous uploads
    for old_file in session_folder.iterdir():
        old_file.unlink()
    
    uploaded_files = []
    csv_count = 0
    txt_count = 0
    
    for file in files:
        if file.filename:
            # Sanitize filename
            filename = Path(file.filename).name
            if filename.lower().endswith('.csv'):
                csv_count += 1
            elif filename.lower().endswith('.txt'):
                txt_count += 1
            else:
                continue  # Skip non-csv/txt files
            
            file_path = session_folder / filename
            file.save(str(file_path))
            uploaded_files.append(filename)
    
    if not uploaded_files:
        return jsonify({'error': 'No valid CSV or TXT files found'}), 400
    
    # Detect unit from files
    unit = get_unit_from_files(str(session_folder))
    
    # Scan for measurement types in text files
    txt_files = list(session_folder.glob('*.txt'))
    file_measurement_types = {}
    
    for txt_file in txt_files:
        types = scan_text_file_for_measurement_types(txt_file)
        if types and len(types) > 1:
            file_measurement_types[txt_file.name] = sorted(list(types))
    
    # Extract test value configurations
    test_configs = extract_test_configs(session_folder, unit)
    
    # Store in session
    session['files_info'] = {
        'count': len(uploaded_files),
        'csv_count': csv_count,
        'txt_count': txt_count,
        'unit': unit,
        'filenames': uploaded_files
    }
    session['measurement_types'] = file_measurement_types
    session['test_configs'] = test_configs
    
    return jsonify({
        'success': True,
        'files_count': len(uploaded_files),
        'csv_count': csv_count,
        'txt_count': txt_count,
        'unit': unit,
        'measurement_types': file_measurement_types,
        'test_configs': test_configs
    })


def extract_test_configs(input_dir, unit):
    """Extract unique (test_value, range_setting, io_type) tuples from filenames."""
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    test_configs = []
    seen = set()
    
    # CSV files are Output devices
    for csv_file in csv_files:
        value, _, channel, range_setting = parse_filename(csv_file.name)
        if value is not None and channel is not None:
            key = (value, range_setting, 'Output')
            if key not in seen:
                seen.add(key)
                test_configs.append({
                    'test_value': value,
                    'range_setting': range_setting if range_setting else 'N/A',
                    'io_type': 'Output',
                    'reference': value,
                    'tolerance': 0.015
                })
    
    # TXT files are Input devices
    for txt_file in txt_files:
        value, _, _, range_setting = parse_filename(txt_file.name)
        if value is not None:
            key = (value, range_setting, 'Input')
            if key not in seen:
                seen.add(key)
                test_configs.append({
                    'test_value': value,
                    'range_setting': range_setting if range_setting else 'N/A',
                    'io_type': 'Input',
                    'reference': value,
                    'tolerance': 0.015
                })
    
    # Sort by test value, then I/O type, then range
    test_configs.sort(key=lambda x: (x['test_value'], x['io_type'], x['range_setting']))
    
    return test_configs


@app.route('/configure')
def configure():
    """Configuration page for measurement types and tolerances."""
    if 'files_info' not in session:
        return redirect(url_for('equipment_report'))
    
    return render_template('configure.html',
                          files_info=session.get('files_info'),
                          measurement_types=session.get('measurement_types', {}),
                          test_configs=session.get('test_configs', []))


@app.route('/api/process', methods=['POST'])
def process_files():
    """Process uploaded files with user configuration."""
    if 'files_info' not in session:
        return jsonify({'error': 'No files uploaded'}), 400
    
    data = request.json
    measurement_type_selections = data.get('measurement_types', {})
    user_configs = data.get('configs', [])
    
    session_folder = get_session_folder()
    output_folder = get_output_folder()
    
    # Convert user configs to the expected format
    user_inputs = {}
    for config in user_configs:
        test_value = config['test_value']
        range_setting = config['range_setting'] if config['range_setting'] != 'N/A' else None
        io_type = config['io_type']
        
        range_input = config.get('range_input', config['range_setting'])
        final_range = None if range_input == 'N/A' or range_input == '' else range_input
        
        user_inputs[(test_value, range_setting, io_type)] = {
            'range': final_range,
            'reference': float(config['reference']),
            'tolerance': float(config['tolerance'])
        }
    
    # Convert measurement type selections keys
    full_path_selections = {}
    for filename, selected_type in measurement_type_selections.items():
        full_path = str(session_folder / filename)
        full_path_selections[full_path] = selected_type
    
    unit = session['files_info']['unit']
    
    try:
        output_file, html_file = process_measurement_files(
            input_dir=str(session_folder),
            output_dir=str(output_folder),
            user_inputs=user_inputs,
            unit=unit,
            measurement_type_selections=full_path_selections
        )
        
        # Store output file paths in session
        session['output_files'] = {
            'excel': os.path.basename(output_file) if output_file else None,
            'html': os.path.basename(html_file) if html_file else None
        }
        
        return jsonify({
            'success': True,
            'excel_file': os.path.basename(output_file) if output_file else None,
            'html_file': os.path.basename(html_file) if html_file else None
        })
        
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500


def process_measurement_files(input_dir, output_dir, user_inputs, unit, measurement_type_selections=None):
    """
    Process all CSV and TXT files in the input directory and compile results into Excel.
    Modified version of the original process_files function to support web output.
    """
    dir_name = Path(input_dir).name
    if not dir_name:
        dir_name = Path(input_dir).resolve().name
    
    # Generate versioned output filename in output directory
    base_output_file = os.path.join(output_dir, f'{dir_name}.xlsx')
    output_file = get_versioned_filename(base_output_file)
    
    results = []
    
    csv_files = list(Path(input_dir).glob('*.csv'))
    txt_files = list(Path(input_dir).glob('*.txt'))
    
    total_files = len(csv_files) + len(txt_files)
    if total_files == 0:
        raise ValueError(f"No CSV or TXT files found in {input_dir}")
    
    # Get timestamp of the first data file (for report header)
    all_data_files = csv_files + txt_files
    if all_data_files:
        first_file = min(all_data_files, key=lambda f: f.stat().st_mtime)
        data_file_timestamp = datetime.fromtimestamp(first_file.stat().st_mtime)
    else:
        data_file_timestamp = None
    
    # Process CSV files (output data)
    for csv_file in csv_files:
        value, file_unit, channel, range_setting = parse_filename(csv_file.name)
        
        if value is None or channel is None:
            continue
        
        try:
            df = pd.read_csv(csv_file)
            
            measurement_col = None
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['voltage', 'vdc', 'resistance', 'ohm', 'current', 'adc', 'measurement']):
                    measurement_col = col
                    break
            
            if measurement_col is None:
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) == 0:
                    continue
                measurement_col = numeric_cols[-1]
            
            measurements = df[measurement_col].dropna()
            
            if len(measurements) == 0:
                continue
            
            result = {
                'Channel': channel,
                'I/O Type': 'Output',
                'Range Setting': range_setting if range_setting else 'N/A',
                f'Test Value [{unit}]': value,
                f'Mean [{unit}]': measurements.mean(),
                f'StdDev [{unit}]': measurements.std(),
                f'Min [{unit}]': measurements.min(),
                f'Max [{unit}]': measurements.max(),
                'Samples': len(measurements),
                '_range_key': range_setting
            }
            
            results.append(result)
            
        except Exception as e:
            continue
    
    # Process TXT files (input data)
    for txt_file in txt_files:
        value, file_unit, channel_from_name, range_setting = parse_filename(txt_file.name)
        
        if value is None:
            continue
        
        try:
            # Get selected measurement type for this file
            selected_type = None
            if measurement_type_selections and str(txt_file) in measurement_type_selections:
                selected_type = measurement_type_selections[str(txt_file)]
            
            # Parse the text file
            channel_data = parse_text_file(txt_file, selected_measurement_type=selected_type,
                                          channel_from_filename=channel_from_name)
            
            if not channel_data:
                continue
            
            # Process each channel found in the file
            for channel, measurements in channel_data.items():
                if len(measurements) == 0:
                    continue
                
                measurements = pd.Series(measurements).dropna()
                
                if len(measurements) == 0:
                    continue
                
                result = {
                    'Channel': channel,
                    'I/O Type': 'Input',
                    'Range Setting': range_setting if range_setting else 'N/A',
                    f'Test Value [{unit}]': value,
                    f'Mean [{unit}]': measurements.mean(),
                    f'StdDev [{unit}]': measurements.std(),
                    f'Min [{unit}]': measurements.min(),
                    f'Max [{unit}]': measurements.max(),
                    'Samples': len(measurements),
                    '_range_key': range_setting
                }
                
                results.append(result)
            
        except Exception as e:
            continue
    
    if not results:
        raise ValueError("No valid results to save")
    
    # Create DataFrame and sort
    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values(['Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]'])
    
    # Add reference value, tolerance, and limits columns
    if user_inputs:
        def get_user_config(row):
            test_val = row[f'Test Value [{unit}]']
            range_key = row['_range_key']
            io_type = row['I/O Type']
            key = (test_val, range_key, io_type)
            return user_inputs.get(key, {})
        
        df_results['_config'] = df_results.apply(get_user_config, axis=1)
        
        df_results[f'Reference Value [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('reference', np.nan) if x else np.nan
        )
        df_results[f'Tolerance [{unit}]'] = df_results['_config'].apply(
            lambda x: x.get('tolerance', np.nan) if x else np.nan
        )
        
        df_results['Range Setting'] = df_results.apply(
            lambda row: row['_config'].get('range', row['Range Setting'])
                        if row['_config'] and row['_config'].get('range') is not None
                        else row['Range Setting'] if row['Range Setting'] != 'N/A' else 'N/A',
            axis=1
        )
        
        # Calculate limits using reference value
        df_results[f'Lower Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] - df_results[f'Tolerance [{unit}]']
        df_results[f'Upper Limit [{unit}]'] = df_results[f'Reference Value [{unit}]'] + df_results[f'Tolerance [{unit}]']
        
        df_results['Mean Check'] = df_results.apply(
            lambda row: 'PASS' if row[f'Lower Limit [{unit}]'] <= row[f'Mean [{unit}]'] <= row[f'Upper Limit [{unit}]'] else 'FAIL',
            axis=1
        )
        
        df_results['Mean±2σ Check'] = df_results.apply(
            lambda row: 'PASS' if (
                row[f'Lower Limit [{unit}]'] <= (row[f'Mean [{unit}]'] - 2*row[f'StdDev [{unit}]']) and
                (row[f'Mean [{unit}]'] + 2*row[f'StdDev [{unit}]']) <= row[f'Upper Limit [{unit}]']
            ) else 'FAIL',
            axis=1
        )
        
        df_results = df_results.drop(columns=['_config', '_range_key'])
        
        column_order = [
            'Channel', 'I/O Type', 'Range Setting', f'Test Value [{unit}]',
            f'Reference Value [{unit}]', f'Tolerance [{unit}]',
            f'Lower Limit [{unit}]', f'Upper Limit [{unit}]',
            f'Mean [{unit}]', f'StdDev [{unit}]', f'Min [{unit}]', f'Max [{unit}]',
            'Samples', 'Mean Check', 'Mean±2σ Check'
        ]
        df_results = df_results[column_order]
    else:
        df_results = df_results.drop(columns=['_range_key'])
    
    # Save to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_results.to_excel(writer, sheet_name='Test Results', index=False)
        
        worksheet = writer.sheets['Test Results']
        worksheet.auto_filter.ref = worksheet.dimensions
        
        if user_inputs:
            numeric_cols = [4, 5, 6, 7, 8, 9, 10, 11, 12]
            samples_col = 13
            pass_fail_cols = [14, 15]
        else:
            numeric_cols = [4, 5, 6, 7, 8]
            samples_col = 9
            pass_fail_cols = []
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if cell.column in numeric_cols:
                    cell.number_format = '0.000000'
                elif cell.column == samples_col:
                    cell.number_format = '0'
                elif cell.column in pass_fail_cols:
                    if cell.value == 'PASS':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100', bold=True)
                    elif cell.value == 'FAIL':
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                        cell.font = Font(color='9C0006', bold=True)
        
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    html_file = None
    if user_inputs:
        color_assignments = create_tolerance_charts(output_file, df_results, unit)
        
        # Apply channel colors to Test Results sheet
        if color_assignments:
            apply_channel_colors_to_results(output_file, df_results, unit, color_assignments)
        
        # Generate interactive HTML report
        html_file = create_html_report(output_file, df_results, unit, data_file_timestamp)
    
    return output_file, html_file


@app.route('/results')
def results():
    """Results page showing generated reports."""
    if 'output_files' not in session:
        return redirect(url_for('equipment_report'))
    
    return render_template('results.html',
                          output_files=session.get('output_files'))


@app.route('/download/<filename>')
def download_file(filename):
    """Download generated files."""
    output_folder = get_output_folder()
    file_path = output_folder / filename
    
    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(str(file_path), as_attachment=True)


@app.route('/view/<filename>')
def view_file(filename):
    """View HTML report in browser."""
    output_folder = get_output_folder()
    file_path = output_folder / filename
    
    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(str(file_path))


@app.route('/api/save-config', methods=['POST'])
def save_config():
    """Save configuration to JSON file."""
    data = request.json
    output_folder = get_output_folder()
    
    config_file = output_folder / 'test_config.json'
    
    with open(config_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)
    
    return jsonify({
        'success': True,
        'filename': 'test_config.json'
    })


@app.route('/api/load-config', methods=['POST'])
def load_config():
    """Load configuration from uploaded JSON file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    try:
        content = file.read().decode('utf-8')
        config_data = json.loads(content)
        return jsonify({
            'success': True,
            'config': config_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/reset')
def reset_session():
    """Reset session and clean up files."""
    if 'session_id' in session:
        session_folder = UPLOAD_FOLDER / session['session_id']
        output_folder = OUTPUT_FOLDER / session['session_id']
        
        if session_folder.exists():
            shutil.rmtree(session_folder)
        if output_folder.exists():
            shutil.rmtree(output_folder)
    
    session.clear()
    return jsonify({'success': True})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
